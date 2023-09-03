"""
Script Name: Keyword Replacement in Excel
Description: This script will check which polygon a point intersects in multiple feature classes and allows you to map the results to keywords in an Excel Spreadsheet. The inputs should be 1) the document filepath including the .xlsx extension with keywords already saved in place, 2) the location of the point to check for intersection, and 3) a LayersToCheck table. For each layer, the LayersToCheck value should be the polygon features, the Attribute value should be the name of the attribute to be returned, and the Keyword that attribute value will replace in the Word Document. Keywords need to be unique, so a recommended format is <<KEYWORD>>. They also cannot include spaces or semicolons. The script is currently set up to only locate the first intersecting polygon--best used for sutations like zip codes or counties where only one value would be returned for a single point.
The OPENPYXL module must be imported prior to running.
Default Notes: Once the LayersToCheck table has been set and run the first time, the full value will print to the results window and can be copied into the default parameter for repeated use.
Tested in ArcGIS Pro 3.1.0
Last Updated: 9/2/2023
"""
import arcpy
import os
import openpyxl
def script_tool(param0, param1, param2):
	
	def CheckIntersections(point_geometry, polygon_feature_class, return_attribute):
		with arcpy.da.SearchCursor(point_geometry, ["SHAPE@"]) as point_cursor:
			for point_row in point_cursor:
				point_geometry = point_row[0]
			with arcpy.da.SearchCursor(polygon_feature_class, [return_attribute,"SHAPE@"]) as cursor:
				for row in cursor:
					polygon_geometry = row[1]
					if polygon_geometry.overlaps(point_geometry) or polygon_geometry.contains(point_geometry):
						intersect_value = row[0]
						break
				else:
					arcpy.AddMessage(f"Point does not intersect with {polygon_feature_class}")
		return intersect_value
	
	point_input=param1
	dict_keywords = dict()
	
	
	#First split string along semicolons for each feature layer
	param2_list = param2.split(";")
	for x in param2_list:
		featurelayer_to_check = x.split(" ")
		p_f_c = featurelayer_to_check[0]
		ret_att = featurelayer_to_check[1]
		new_key = featurelayer_to_check[2]
		result = CheckIntersections(point_input, p_f_c, ret_att)
		dict_keywords[new_key] = str(result)
	
	
	
	#REPLACE TEXT FUNCTION
	workbook = openpyxl.load_workbook(param0)
	worksheet = workbook.worksheets[0]
	xls_path = os.path.dirname(param0)  # Replace with your document's path
	xls_name = os.path.basename(param0)
	xls_name_ext = os.path.splitext(xls_name)
	xlsx_upd_name = xls_name_ext[0] + "_Update" + xls_name_ext[1]
	
	def replace_text_in_xls(dict):
	
		for key, value in dict.items():
			old_text = key
			new_text = value
			for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
				for cell in row:
					if old_text in str(cell.value):
						cell.value = str(cell.value).replace(old_text, new_text)
	
		workbook.save(os.path.join(xls_path,xlsx_upd_name))
	replace_text_in_xls(dict_keywords)
	
	return
if __name__ == "__main__":
	param0 = arcpy.GetParameterAsText(0)
	param1 = arcpy.GetParameterAsText(1)
	param2 = arcpy.GetParameterAsText(2)
	arcpy.AddMessage(param2)
      
	script_tool(param0, param1, param2)
